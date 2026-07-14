from __future__ import annotations

from copy import copy
from datetime import date, datetime
from pathlib import Path
import sys
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .db import Database
from .statistics import StatisticsService


def _write_rows(sheet: Worksheet, headers: list[str], rows: Iterable[dict[str, object]]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)


def _kind_label(kind: str) -> str:
    return "Check" if kind == "CHECK" else "Scan"


class ReportService:
    def __init__(self, db: Database):
        self.db = db

    def export_daily_report(self, project_id: int, output_dir: Path | None = None) -> Path:
        project = self.db.get_project(project_id)
        if not project:
            raise ValueError(f"Project not found: {project_id}")
        output_dir = output_dir or Path(project.reports_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        report_rows = self.db.list_report_rows(project_id)

        summary = workbook.active
        summary.title = "Summary"
        counts = self.db.dashboard_counts(project_id)
        summary.append(["Metric", "Value"])
        for cell in summary[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for key in sorted(counts):
            summary.append([key, counts[key]])
        summary.append(["project_code", project.project_code])
        summary.append(["project_name", project.display_name])
        summary.column_dimensions["A"].width = 28
        summary.column_dimensions["B"].width = 16

        backups = workbook.create_sheet("Backup Files")
        _write_rows(
            backups,
            [
                "id",
                "client_code",
                "project_code",
                "relative_project_path",
                "dest_path",
                "file_size",
                "status",
                "error_message",
                "created_at",
                "copied_at",
                "verified_at",
            ],
            [dict(row) for row in report_rows["backup_files"]],
        )

        conflicts = workbook.create_sheet("Conflicts")
        _write_rows(
            conflicts,
            [
                "id",
                "client_code",
                "source_path",
                "dest_path",
                "status",
                "resolution",
                "archive_path",
                "created_at",
                "resolved_at",
            ],
            [dict(row) for row in report_rows["conflicts"]],
        )

        mapfile = workbook.create_sheet("Mapfile")
        _write_rows(
            mapfile,
            [
                "id", "import_id", "row_number", "expected_relative_path", "status", "message",
                "is_done", "done_at", "done_by",
            ],
            [dict(row) for row in report_rows["mapfile_rows"]],
        )

        personnel = workbook.create_sheet("Personnel")
        _write_rows(
            personnel,
            ["id", "personnel_code", "full_name", "role_name", "enabled", "created_at", "updated_at"],
            [dict(row) for row in report_rows["personnel"]],
        )

        tasks = workbook.create_sheet("Tasks")
        _write_rows(
            tasks,
            [
                "id", "task_code", "title", "description", "personnel_code",
                "assignee_name", "due_date", "priority", "status", "created_at", "updated_at",
            ],
            [dict(row) for row in report_rows["tasks"]],
        )

        output = output_dir / f"scan_backup_report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        workbook.save(output)
        self.db.record_audit(
            "REPORT_EXPORTED", f"Exported report to {output}", project_id=project_id
        )
        return output

    def export_statistics_report(
        self, project_id: int, date_from: str, date_to: str, output_dir: Path | None = None
    ) -> Path:
        project = self.db.get_project(project_id)
        if not project:
            raise ValueError(f"Project not found: {project_id}")
        output_dir = output_dir or Path(project.reports_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        stats = StatisticsService(self.db)
        workbook = Workbook()

        daily = workbook.active
        daily.title = "Daily"
        _write_rows(
            daily,
            ["day", "done_count", "backed_up_count"],
            [
                {"day": row.day, "done_count": row.done_count, "backed_up_count": row.backed_up_count}
                for row in stats.productivity_by_day(project_id, date_from, date_to)
            ],
        )

        personnel = workbook.create_sheet("Personnel")
        _write_rows(
            personnel,
            ["personnel_code", "full_name", "done_count"],
            [
                {
                    "personnel_code": row.personnel_code,
                    "full_name": row.full_name,
                    "done_count": row.done_count,
                }
                for row in stats.productivity_by_personnel(project_id, date_from, date_to)
            ],
        )

        paper_sizes = workbook.create_sheet("Paper Sizes")
        _write_rows(
            paper_sizes,
            ["paper_code", "page_count", "file_count"],
            [
                {
                    "paper_code": row.paper_code,
                    "page_count": row.page_count,
                    "file_count": row.file_count,
                }
                for row in stats.paper_size_summary(project_id, date_from, date_to)
            ],
        )

        ratio = stats.completion_ratio(project_id)
        latency = stats.done_to_backup_latency(project_id, date_from, date_to)
        summary = workbook.create_sheet("Summary")
        summary.append(["Metric", "Value"])
        for cell in summary[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        summary.append(["date_from", date_from])
        summary.append(["date_to", date_to])
        summary.append(["total_rows", ratio.total_rows])
        summary.append(["done_count", ratio.done_count])
        summary.append(["matched_count", ratio.matched_count])
        summary.append(["done_pct", round(ratio.done_pct, 1)])
        summary.append(["matched_pct", round(ratio.matched_pct, 1)])
        summary.append(["latency_sample_count", latency.sample_count])
        summary.append([
            "latency_average_hours",
            round(latency.average_hours, 2) if latency.average_hours is not None else "",
        ])
        summary.append([
            "latency_median_hours",
            round(latency.median_hours, 2) if latency.median_hours is not None else "",
        ])
        summary.append(["latency_bucket_under_1h", latency.bucket_under_1h])
        summary.append(["latency_bucket_1_to_4h", latency.bucket_1_to_4h])
        summary.append(["latency_bucket_4_to_24h", latency.bucket_4_to_24h])
        summary.append(["latency_bucket_over_24h", latency.bucket_over_24h])
        summary.column_dimensions["A"].width = 28
        summary.column_dimensions["B"].width = 16

        output = (
            output_dir
            / f"statistics_report_{date_from}_{date_to}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        workbook.save(output)
        self.db.record_audit(
            "STATISTICS_REPORT_EXPORTED", f"Exported statistics report to {output}",
            project_id=project_id,
        )
        return output

    def export_attendance_report(
        self, project_id: int, date_from: str, date_to: str, output_dir: Path | None = None
    ) -> Path:
        project = self.db.get_project(project_id)
        if not project:
            raise ValueError(f"Project not found: {project_id}")
        output_dir = output_dir or Path(project.reports_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        stats = StatisticsService(self.db)
        details = stats.personnel_daily_job_details(project_id, date_from, date_to)
        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Cham cong"
        _write_rows(
            sheet,
            [
                "day",
                "personnel_code",
                "full_name",
                "sequence_number",
                "job_title",
                "task_kind",
                "quantity",
                "completed_count",
                "started_at",
                "last_updated_at",
            ],
            [
                {
                    "day": row.day,
                    "personnel_code": row.personnel_code,
                    "full_name": row.full_name,
                    "sequence_number": row.sequence_number,
                    "job_title": row.job_title,
                    "task_kind": _kind_label(row.task_kind),
                    "quantity": row.quantity,
                    "completed_count": row.completed_count,
                    "started_at": row.started_at,
                    "last_updated_at": row.last_updated_at,
                }
                for row in details
            ],
        )

        summary: dict[tuple[str, str, str], dict[str, object]] = {}
        for row in details:
            key = (row.day, row.personnel_code, row.full_name)
            bucket = summary.setdefault(
                key,
                {
                    "day": row.day,
                    "personnel_code": row.personnel_code,
                    "full_name": row.full_name,
                    "job_count": 0,
                    "total_quantity": 0,
                    "completed_count": 0,
                    "first_started_at": row.started_at,
                },
            )
            bucket["job_count"] = int(bucket["job_count"]) + 1
            bucket["total_quantity"] = int(bucket["total_quantity"]) + row.quantity
            bucket["completed_count"] = int(bucket["completed_count"]) + row.completed_count
            first_started = str(bucket["first_started_at"] or "")
            if row.started_at and (not first_started or row.started_at < first_started):
                bucket["first_started_at"] = row.started_at

        summary_sheet = workbook.create_sheet("Tong hop")
        _write_rows(
            summary_sheet,
            [
                "day",
                "personnel_code",
                "full_name",
                "job_count",
                "total_quantity",
                "completed_count",
                "first_started_at",
            ],
            [
                summary[key]
                for key in sorted(summary, key=lambda item: (item[0], item[2], item[1]))
            ],
        )

        raw_entries = self.db.list_attendance_entries(project_id, date_from, date_to)
        raw_sheet = workbook.create_sheet("San luong tho")
        _write_rows(
            raw_sheet,
            [
                "id",
                "work_date",
                "personnel_code",
                "full_name",
                "record_key",
                "job_title",
                "task_kind",
                "quantity",
                "completed_count",
                "status",
                "task_status",
                "record_status",
                "approved_by",
                "approved_at",
                "override_reason",
                "notes",
            ],
            [
                {
                    "id": row["id"],
                    "work_date": row["work_date"],
                    "personnel_code": row["personnel_code"],
                    "full_name": row["full_name"],
                    "record_key": row["record_key"],
                    "job_title": row["job_title"],
                    "task_kind": _kind_label(row["task_kind"]),
                    "quantity": row["quantity"],
                    "completed_count": row["completed_count"],
                    "status": row["status"],
                    "task_status": row["task_status"] or "",
                    "record_status": row["record_status"],
                    "approved_by": row["approved_by"],
                    "approved_at": row["approved_at"],
                    "override_reason": row["override_reason"],
                    "notes": row["notes"],
                }
                for row in raw_entries
            ],
        )

        exception_sheet = workbook.create_sheet("Ngoai le")
        _write_rows(
            exception_sheet,
            [
                "id",
                "work_date",
                "personnel_code",
                "full_name",
                "record_key",
                "job_title",
                "task_kind",
                "status",
                "task_status",
                "record_status",
                "has_scan_backup",
                "check_pages",
                "check_files",
                "notes",
            ],
            [
                {
                    "id": row["id"],
                    "work_date": row["work_date"],
                    "personnel_code": row["personnel_code"],
                    "full_name": row["full_name"],
                    "record_key": row["record_key"],
                    "job_title": row["job_title"],
                    "task_kind": _kind_label(row["task_kind"]),
                    "status": row["status"],
                    "task_status": row["task_status"] or "",
                    "record_status": row["record_status"],
                    "has_scan_backup": int(row["has_scan_backup"] or 0),
                    "check_pages": row["check_pages"],
                    "check_files": row["check_files"],
                    "notes": row["notes"],
                }
                for row in raw_entries
                if row["status"] != "APPROVED"
                or row["override_reason"]
                or (row["task_kind"] == "SCAN" and not row["has_scan_backup"])
                or (row["task_kind"] == "CHECK" and row["record_status"] != "COMPLETED")
            ],
        )

        with self.db.connect() as conn:
            audit_rows = conn.execute(
                """
                SELECT action, message, created_at FROM audit_logs
                WHERE project_id=?
                    AND action IN (
                        'ATTENDANCE_APPROVED',
                        'ATTENDANCE_REJECTED',
                        'ATTENDANCE_REPORT_EXPORTED'
                    )
                    AND substr(created_at, 1, 10) BETWEEN ? AND ?
                ORDER BY created_at DESC
                """,
                (project_id, date_from, date_to),
            ).fetchall()
        audit_sheet = workbook.create_sheet("Audit chinh sua")
        _write_rows(
            audit_sheet,
            ["created_at", "action", "message"],
            [
                {
                    "created_at": row["created_at"],
                    "action": row["action"],
                    "message": row["message"],
                }
                for row in audit_rows
            ],
        )

        output = (
            output_dir
            / f"attendance_report_{date_from}_{date_to}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        workbook.save(output)
        self.db.record_audit(
            "ATTENDANCE_REPORT_EXPORTED",
            f"Exported attendance report to {output}",
            project_id=project_id,
        )
        return output

    def export_mausham_cong(
        self, project_id: int, date_from: str, date_to: str, output_dir: Path | None = None
    ) -> Path:
        """Export the official timesheet in the MauChamCong.xlsx layout: one
        sheet per work day, each person a numbered 4-row block (4 job slots)
        with name, hours, attendance type, job content and output volume.

        Only APPROVED attendance counts — the leader signs off in the Workbench
        first."""
        project = self.db.get_project(project_id)
        if not project:
            raise ValueError(f"Project not found: {project_id}")
        output_dir = output_dir or Path(project.reports_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        entries = self.db.list_attendance_entries(
            project_id, date_from, date_to, statuses=["APPROVED"]
        )
        by_date: dict[str, dict[int, list]] = {}
        person_name: dict[int, tuple[str, str]] = {}
        for row in entries:
            person_name[row["personnel_id"]] = (row["personnel_code"], row["full_name"])
            by_date.setdefault(row["work_date"], {}).setdefault(row["personnel_id"], []).append(row)

        work_dates = sorted(by_date) if by_date else [date_from]
        workbook = _load_mausham_template()
        template_sheet = workbook.active
        sheets = [template_sheet]
        sheets.extend(workbook.copy_worksheet(template_sheet) for _ in work_dates[1:])
        for sheet, work_date in zip(sheets, work_dates, strict=True):
            sheet.title = work_date[:31]
            _build_mausham_sheet(sheet, work_date, by_date.get(work_date, {}), person_name)

        output = (
            output_dir
            / f"MauChamCong_{date_from}_{date_to}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        workbook.save(output)
        self.db.record_audit(
            "MAUSHAMCONG_EXPORTED",
            f"Exported MauChamCong timesheet to {output}",
            project_id=project_id,
        )
        return output


_WEEKDAYS_VI = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
_MAUSHAM_HEADERS = (
    "STT",
    "Họ Và Tên",
    "Số lượng công việc thực hiện trong ngày",
    "Thời gian thực hiện từng mục công việc",
    "Loại Chấm Công/Năng Suất",
    "Nội dung công việc",
    "Khối lượng hoàn thành",
)
_MAUSHAM_JOB_SLOTS = 4
_MAUSHAM_TEMPLATE_NAME = "MauChamCong.xlsx"
_MAUSHAM_TITLE_MERGES = {"A1:C2", "D1:G1", "D2:G2"}


def _weekday_vi(work_date: str) -> str:
    try:
        return _WEEKDAYS_VI[date.fromisoformat(work_date).weekday()]
    except (ValueError, TypeError):
        return ""


def _mausham_template_path() -> Path:
    candidates: list[Path] = []
    bundle_dir = getattr(sys, "_MEIPASS", None)
    if bundle_dir:
        candidates.append(Path(bundle_dir) / _MAUSHAM_TEMPLATE_NAME)
    candidates.extend(
        [
            Path(__file__).resolve().parents[2] / _MAUSHAM_TEMPLATE_NAME,
            Path(sys.executable).resolve().parent / _MAUSHAM_TEMPLATE_NAME,
            Path.cwd() / _MAUSHAM_TEMPLATE_NAME,
        ]
    )
    for candidate in dict.fromkeys(candidates):
        if candidate.is_file():
            return candidate
    searched = ", ".join(str(path) for path in dict.fromkeys(candidates))
    raise FileNotFoundError(f"Không tìm thấy file mẫu {_MAUSHAM_TEMPLATE_NAME}. Đã kiểm tra: {searched}")


def _load_mausham_template():
    workbook = load_workbook(_mausham_template_path())
    if len(workbook.worksheets) != 1:
        raise ValueError(f"{_MAUSHAM_TEMPLATE_NAME} phải có đúng một sheet mẫu.")
    sheet = workbook.active
    headers = tuple(sheet.cell(row=3, column=column).value for column in range(1, 8))
    merges = {str(merged) for merged in sheet.merged_cells.ranges}
    if headers != _MAUSHAM_HEADERS or not _MAUSHAM_TITLE_MERGES.issubset(merges):
        raise ValueError(f"{_MAUSHAM_TEMPLATE_NAME} không đúng cấu trúc biểu mẫu chấm công.")
    if sheet.max_column != len(_MAUSHAM_HEADERS):
        raise ValueError(f"{_MAUSHAM_TEMPLATE_NAME} chỉ được có 7 cột từ A đến G.")
    return workbook


def _build_mausham_sheet(
    sheet: Worksheet,
    work_date: str,
    people: dict[int, list],
    person_name: dict[int, tuple[str, str]],
) -> None:
    try:
        parsed_date = date.fromisoformat(work_date)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Ngày chấm công không hợp lệ: {work_date}") from exc

    body_styles = {
        "index": copy(sheet["A4"]._style),
        "text": copy(sheet["B4"]._style),
        "job": copy(sheet["C4"]._style),
        "number": copy(sheet["D4"]._style),
    }
    body_row_height = sheet.row_dimensions[4].height or 15

    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= 4:
            sheet.unmerge_cells(str(merged))
    if sheet.max_row > 3:
        sheet.delete_rows(4, sheet.max_row - 3)

    sheet["D1"] = datetime(parsed_date.year, parsed_date.month, parsed_date.day)
    sheet["D2"] = _weekday_vi(work_date)

    row = 4
    ordered_people = sorted(
        people,
        key=lambda personnel_id: (
            person_name.get(personnel_id, ("", ""))[1].casefold(),
            person_name.get(personnel_id, ("", ""))[0].casefold(),
        ),
    )
    for sequence, personnel_id in enumerate(ordered_people, start=1):
        jobs = _group_jobs_by_type(people[personnel_id])
        _code, name = person_name.get(personnel_id, ("", ""))
        top, bottom = row, row + _MAUSHAM_JOB_SLOTS - 1
        for slot in range(_MAUSHAM_JOB_SLOTS):
            r = top + slot
            sheet.row_dimensions[r].height = body_row_height
            for column in range(1, 8):
                cell = sheet.cell(row=r, column=column)
                cell._style = copy(
                    body_styles["job"] if column == 3 else body_styles["text"]
                )
            sheet.cell(row=r, column=3, value=f"Công việc {slot + 1}")
            job = jobs[slot] if slot < len(jobs) else None
            if job is not None:
                hours = job["hours"] or 0
                if hours:
                    sheet.cell(row=r, column=4, value=hours)._style = copy(
                        body_styles["number"]
                    )
                sheet.cell(row=r, column=5, value=job["attendance_type"] or None)
                sheet.cell(row=r, column=6, value=job["name"])
                if job["volume"]:
                    sheet.cell(row=r, column=7, value=job["volume"])._style = copy(
                        body_styles["number"]
                    )
        sheet[f"A{top}"]._style = copy(body_styles["index"])
        sheet[f"A{top}"] = sequence
        sheet[f"B{top}"] = name
        sheet.merge_cells(f"A{top}:A{bottom}")
        row = bottom + 1


def _group_jobs_by_type(entries: list) -> list[dict]:
    """Collapse a person's attendance rows into one entry per distinct job type
    (MauChamCong "Công việc" slot): the output volume is summed across every
    record of the job type, hours/type come from the first row (all rows of a
    slot share them after the leader saves)."""
    grouped: dict[str, dict] = {}
    order: list[str] = []
    for entry in entries:
        name = (entry["job_content"] or entry["job_title"] or "").strip()
        key = (entry["job_title"] or entry["job_content"] or "").strip().casefold()
        if key not in grouped:
            grouped[key] = {
                "name": name,
                "hours": entry["work_hours"] or 0,
                "attendance_type": entry["attendance_type"] or None,
                "volume": int(entry["quantity"] or 0),
            }
            order.append(key)
        else:
            grouped[key]["volume"] += int(entry["quantity"] or 0)
            if not grouped[key]["hours"] and (entry["work_hours"] or 0):
                grouped[key]["hours"] = entry["work_hours"]
            if not grouped[key]["attendance_type"] and entry["attendance_type"]:
                grouped[key]["attendance_type"] = entry["attendance_type"]
    return [grouped[key] for key in order]
