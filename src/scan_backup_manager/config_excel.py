from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .db import Database
from .models import Client, Personnel


CLIENT_HEADERS = ["client_code", "share_path", "enabled", "notes"]
PERSONNEL_HEADERS = ["personnel_code", "full_name", "role_name", "enabled", "pin"]


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_bool(value: Any, *, default: bool = True) -> bool:
    text = _normalize_cell(value).strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "y", "co", "có", "bat", "bật", "active", "enabled"}


def _read_rows(path: Path, required_headers: list[str]) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_values:
        raise ValueError("File Excel không có dòng tiêu đề.")
    headers = [_normalize_cell(value) for value in header_values]
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError(f"Thiếu cột Excel: {', '.join(missing)}")
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        raw = {
            header: _normalize_cell(values[index]) if index < len(values) else ""
            for index, header in enumerate(headers)
            if header
        }
        if any(raw.values()):
            rows.append(raw)
    return rows


class ConfigExcelService:
    def __init__(self, db: Database):
        self.db = db

    def export_clients(self, project_id: int, output_dir: Path) -> Path:
        output_dir.mkdir(parents=True, exist_ok=True)
        path = output_dir / "may_tram.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "May tram"
        sheet.append(CLIENT_HEADERS)
        for client in self.db.list_clients(project_id):
            sheet.append([
                client.client_code,
                client.share_path,
                1 if client.enabled else 0,
                client.notes,
            ])
        workbook.save(path)
        return path

    def import_clients(self, project_id: int, path: Path) -> int:
        count = 0
        for row in _read_rows(path, CLIENT_HEADERS):
            code = row.get("client_code", "").strip()
            share_path = row.get("share_path", "").strip()
            if not code or not share_path:
                continue
            self.db.save_client(
                Client(
                    None,
                    project_id,
                    code,
                    "",
                    share_path,
                    _as_bool(row.get("enabled"), default=True),
                    row.get("notes", ""),
                )
            )
            count += 1
        return count

    def export_personnel(self, project_id: int, output_dir: Path) -> Path:
        output_dir.mkdir(parents=True, exist_ok=True)
        path = output_dir / "nhan_su.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Nhan su"
        sheet.append(PERSONNEL_HEADERS)
        for person in self.db.list_personnel(project_id):
            sheet.append([
                person.personnel_code,
                person.full_name,
                person.role_name,
                1 if person.enabled else 0,
                "",
            ])
        workbook.save(path)
        return path

    def import_personnel(self, project_id: int, path: Path) -> int:
        count = 0
        for row in _read_rows(path, PERSONNEL_HEADERS):
            code = row.get("personnel_code", "").strip()
            full_name = row.get("full_name", "").strip()
            if not code or not full_name:
                continue
            personnel_id = self.db.save_personnel(
                Personnel(
                    None,
                    project_id,
                    code,
                    full_name,
                    row.get("role_name", ""),
                    _as_bool(row.get("enabled"), default=True),
                )
            )
            pin = row.get("pin", "").strip()
            if pin:
                self.db.set_personnel_pin(personnel_id, pin, must_change=True)
            count += 1
        return count
