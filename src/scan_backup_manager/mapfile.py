from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .constants import COUNTABLE_BACKUP_STATUSES
from .db import Database
from .filesystem import find_project_roots, workstation_project_root_name
from .models import DirectoryLevel, MapfileProfile


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_expected_path(
    project: str,
    year: str,
    case_type: str,
    case_number: str,
    file_name: str,
) -> str:
    file_name = file_name if file_name.lower().endswith(".pdf") else f"{file_name}.pdf"
    return str(Path(project.strip()) / year.strip() / case_type.strip().upper() / case_number.strip() / file_name.strip())


def normalize_dynamic_expected_path(
    project: str,
    directory_values: list[str],
    file_name: str,
) -> str:
    file_name = file_name.strip()
    file_name = file_name if file_name.lower().endswith(".pdf") else f"{file_name}.pdf"
    return str(Path(project.strip()) / Path(*[value.strip() for value in directory_values]) / file_name)


def safe_folder_name(value: str) -> str:
    clean = str(value).strip() or "-"
    for char in '\\/:*?"<>|':
        clean = clean.replace(char, "-")
    return clean


class MapfileService:
    def __init__(self, db: Database):
        self.db = db

    def import_excel(
        self, project_id: int, file_path: Path, profile: MapfileProfile | None = None
    ) -> int:
        profile = profile or self.db.get_mapfile_profile(project_id)
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook[profile.sheet_name] if profile.sheet_name else workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = {normalize_cell(value): index for index, value in enumerate(header_row)}
        directory_levels = self.db.list_directory_levels(project_id)
        dynamic_level_columns = [level.display_name for level in directory_levels]
        has_dynamic_columns = bool(dynamic_level_columns) and all(
            column in headers for column in dynamic_level_columns
        )
        level_columns = dynamic_level_columns if has_dynamic_columns else [
            profile.year_column,
            profile.case_type_column,
            profile.case_number_column,
        ]
        required = [profile.project_column, *level_columns, profile.file_name_column]
        missing = [column for column in required if column not in headers]
        if missing:
            raise ValueError(f"Missing mapfile columns: {', '.join(missing)}")

        # Carry over the "Done" flag for rows that match a previous import of the
        # same project, so re-importing an updated mapfile doesn't force personnel
        # to re-tick rows they already marked as scanned.
        previous_import_id = self.db.latest_mapfile_import_id(project_id)
        done_by_path: dict[str, tuple[bool, str, int | None]] = {}
        if previous_import_id:
            for previous_row in self.db.list_mapfile_rows(previous_import_id):
                if previous_row["is_done"]:
                    done_by_path[previous_row["expected_relative_path"]] = (
                        True, previous_row["done_at"], previous_row["done_by"],
                    )

        import_id = self.db.create_mapfile_import(project_id, profile.id or 0, str(file_path))
        rows: list[tuple[int, dict[str, Any], str]] = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            raw = {header: normalize_cell(values[index]) for header, index in headers.items() if header}
            if not any(raw.values()):
                continue
            if has_dynamic_columns:
                expected = normalize_dynamic_expected_path(
                    raw.get(profile.project_column, ""),
                    [raw.get(column, "") for column in level_columns],
                    raw.get(profile.file_name_column, ""),
                )
            else:
                expected = normalize_expected_path(
                    raw.get(profile.project_column, ""),
                    raw.get(profile.year_column, ""),
                    raw.get(profile.case_type_column, ""),
                    raw.get(profile.case_number_column, ""),
                    raw.get(profile.file_name_column, ""),
                )
            rows.append((row_number, raw, expected))
        self.db.add_mapfile_rows(import_id, rows)

        if done_by_path:
            for new_row in self.db.list_mapfile_rows(import_id):
                carried = done_by_path.get(new_row["expected_relative_path"])
                if carried:
                    _, done_at, done_by = carried
                    self.db.mark_mapfile_row_done(new_row["id"], done_by, done_at=done_at)

        self.reconcile(project_id, import_id)
        self.db.record_audit(
            "MAPFILE_IMPORTED", f"Imported {len(rows)} rows from {file_path}", project_id=project_id
        )
        return import_id

    def reconcile(self, project_id: int, import_id: int | None = None) -> dict[str, int]:
        import_id = import_id or self.db.latest_mapfile_import_id(project_id)
        if import_id is None:
            return {"matched": 0, "missing": 0, "extra": 0}

        backup_rows = self.db.list_backup_files(project_id, limit=None)
        backed_up = {
            str(Path(row["project_code"]) / row["relative_project_path"]): row
            for row in backup_rows
            if row["status"] in COUNTABLE_BACKUP_STATUSES
        }

        matched = 0
        missing = 0
        for row in self.db.list_mapfile_rows(import_id):
            expected = row["expected_relative_path"]
            if expected in backed_up:
                self.db.update_mapfile_row_status(row["id"], "MATCHED", "Found in backup log")
                matched += 1
            else:
                self.db.update_mapfile_row_status(row["id"], "MISSING", "Not found in backup log")
                missing += 1

        expected_paths = {row["expected_relative_path"] for row in self.db.list_mapfile_rows(import_id)}
        extra = sum(1 for path in backed_up if path not in expected_paths)
        return {"matched": matched, "missing": missing, "extra": extra}

    def reconcile_row(self, project_id: int, row_id: int) -> str:
        """Re-check a single mapfile row against backup_files, without re-running
        reconcile() for the whole import (which can be thousands of rows)."""
        row = self.db.get_mapfile_row(row_id)
        if not row:
            raise ValueError(f"Mapfile row not found: {row_id}")
        matched = self.db.find_backup_file_by_relative_path(project_id, row["expected_relative_path"])
        if matched:
            self.db.update_mapfile_row_status(row_id, "MATCHED", "Found in backup log")
            return "MATCHED"
        self.db.update_mapfile_row_status(row_id, "MISSING", "Not found in backup log")
        return "MISSING"

    def update_row_cell(
        self,
        project_id: int,
        row_id: int,
        column_name: str,
        value: Any,
    ) -> str:
        row = self.db.get_mapfile_row(row_id)
        if not row:
            raise ValueError(f"Mapfile row not found: {row_id}")
        raw = dict(json.loads(row["raw_json"]))
        raw[column_name] = normalize_cell(value)

        profile = self.db.get_mapfile_profile(project_id)
        directory_levels = self.db.list_directory_levels(project_id)
        dynamic_level_columns = [level.display_name for level in directory_levels]
        if dynamic_level_columns and all(column in raw for column in dynamic_level_columns):
            expected = normalize_dynamic_expected_path(
                raw.get(profile.project_column, ""),
                [raw.get(column, "") for column in dynamic_level_columns],
                raw.get(profile.file_name_column, ""),
            )
        else:
            expected = normalize_expected_path(
                raw.get(profile.project_column, ""),
                raw.get(profile.year_column, ""),
                raw.get(profile.case_type_column, ""),
                raw.get(profile.case_number_column, ""),
                raw.get(profile.file_name_column, ""),
            )
        self.db.update_mapfile_row_source(row_id, raw, expected)
        status = self.reconcile_row(project_id, row_id)
        self.db.record_audit(
            "MAPFILE_ROW_UPDATED",
            f"Updated mapfile row {row_id} column {column_name}",
            project_id=project_id,
        )
        return status

    def add_manual_record(
        self,
        project_id: int,
        record_parts: list[str],
        *,
        file_name: str = "",
        client_code: str | None = None,
        workstation_owner: str = "",
        workstation_date: str = "",
        workstation_task: str = "",
    ) -> int:
        project = self.db.get_project(project_id)
        if not project:
            raise ValueError(f"Project not found: {project_id}")
        clean_parts = [str(part).strip() for part in record_parts]
        if not clean_parts or any(not part for part in clean_parts):
            raise ValueError("Cần nhập đầy đủ thông tin hồ sơ.")

        profile = self.db.get_mapfile_profile(project_id)
        directory_levels = self.db.list_directory_levels(project_id)
        self._validate_catalog_required_parts(directory_levels, clean_parts)
        raw: dict[str, Any] = {
            profile.project_column: project.project_code,
            profile.file_name_column: normalize_cell(file_name),
        }
        if clean_parts:
            raw[profile.year_column] = clean_parts[0]
        if len(clean_parts) > 1:
            raw[profile.case_type_column] = clean_parts[1]
        if len(clean_parts) > 2:
            raw[profile.case_number_column] = "/".join(clean_parts[2:])
        for index, value in enumerate(clean_parts):
            if index < len(directory_levels):
                raw[directory_levels[index].display_name] = value

        record_key = "/".join(clean_parts)
        expected_parts = [project.project_code.strip(), *clean_parts]
        if raw[profile.file_name_column]:
            expected_parts.append(raw[profile.file_name_column])
        expected = str(Path(*expected_parts))
        import_id = self.db.latest_mapfile_import_id(project_id)
        if import_id is None:
            import_id = self.db.create_mapfile_import(
                project_id,
                profile.id or 0,
                "manual://system-mapfile",
            )
        if client_code:
            self.create_client_record_folder(
                project_id,
                client_code,
                clean_parts,
                owner_name=workstation_owner,
                work_date=workstation_date,
                task_name=workstation_task,
            )
        row_id = self.db.append_mapfile_row(import_id, raw, expected, record_key=record_key)
        self.reconcile_row(project_id, row_id)
        self.db.record_audit(
            "MAPFILE_ROW_ADDED",
            f"Added manual mapfile row {row_id}",
            project_id=project_id,
        )
        return row_id

    def update_manual_record(
        self,
        project_id: int,
        old_record_key: str,
        record_parts: list[str],
        *,
        file_name: str = "",
    ) -> str:
        project = self.db.get_project(project_id)
        if not project:
            raise ValueError(f"Project not found: {project_id}")
        clean_parts = [str(part).strip() for part in record_parts]
        if not clean_parts or any(not part for part in clean_parts):
            raise ValueError("Cần nhập đầy đủ thông tin hồ sơ.")
        profile = self.db.get_mapfile_profile(project_id)
        directory_levels = self.db.list_directory_levels(project_id)
        self._validate_catalog_required_parts(directory_levels, clean_parts)
        raw: dict[str, Any] = {
            profile.project_column: project.project_code,
            profile.file_name_column: normalize_cell(file_name),
        }
        if clean_parts:
            raw[profile.year_column] = clean_parts[0]
        if len(clean_parts) > 1:
            raw[profile.case_type_column] = clean_parts[1]
        if len(clean_parts) > 2:
            raw[profile.case_number_column] = "/".join(clean_parts[2:])
        for index, value in enumerate(clean_parts):
            if index < len(directory_levels):
                raw[directory_levels[index].display_name] = value

        new_record_key = "/".join(clean_parts)
        expected_parts = [project.project_code.strip(), *clean_parts]
        if raw[profile.file_name_column]:
            expected_parts.append(raw[profile.file_name_column])
        expected = str(Path(*expected_parts))
        self.db.update_system_record_source(
            project_id,
            old_record_key,
            new_record_key,
            raw,
            expected,
        )
        self.db.record_audit(
            "MAPFILE_RECORD_UPDATED",
            f"Updated system mapfile record {old_record_key} -> {new_record_key}",
            project_id=project_id,
        )
        return new_record_key

    @staticmethod
    def _validate_catalog_required_parts(
        directory_levels: list[DirectoryLevel],
        record_parts: list[str],
    ) -> None:
        for index, level in enumerate(directory_levels):
            if not level.require_catalog_selection:
                continue
            if index >= len(record_parts):
                raise ValueError(f"Cần chọn {level.display_name} từ danh mục.")
            allowed = {
                value.upper() if level.validation_type == "ENUM" else value
                for value in level.allowed_values
            }
            value = record_parts[index]
            key = value.upper() if level.validation_type == "ENUM" else value
            if key not in allowed:
                raise ValueError(
                    f"{level.display_name} phải chọn từ danh mục đã cấu hình."
                )

    def create_client_record_folder(
        self,
        project_id: int,
        client_code: str,
        record_parts: list[str],
        *,
        owner_name: str = "",
        work_date: str = "",
        task_name: str = "",
    ) -> Path:
        project = self.db.get_project(project_id)
        if not project:
            raise ValueError(f"Project not found: {project_id}")
        client = next(
            (
                item
                for item in self.db.list_clients(project_id)
                if item.client_code == client_code and item.enabled
            ),
            None,
        )
        if not client:
            raise ValueError("Máy nhận hồ sơ cứng không hợp lệ hoặc đang tắt.")
        workstation_root_name = workstation_project_root_name(project.project_code)
        roots = [
            root
            for root in find_project_roots(Path(client.share_path), project.project_code)
            if root.name.upper() == workstation_root_name.upper()
        ]
        project_root = roots[0] if roots else Path(client.share_path) / workstation_root_name
        common_parts = [
            safe_folder_name(owner_name or "Họ tên"),
            safe_folder_name(work_date or datetime.now().strftime("%d-%m-%Y")),
            safe_folder_name(task_name or "Nội dung công việc"),
        ]
        target = project_root / Path(*common_parts) / Path(*record_parts)
        try:
            target.mkdir(parents=True, exist_ok=True)
        except PermissionError as exc:
            raise ValueError(
                f"Không có quyền tạo thư mục trên máy trạm: {target}. "
                "Vui lòng kiểm tra quyền ghi của tài khoản Windows đang chạy app/service."
            ) from exc
        except OSError as exc:
            raise ValueError(f"Không thể tạo thư mục trên máy trạm: {target}. Lỗi: {exc}") from exc
        return target

    def duplicate_manual_record(self, project_id: int, source_record_key: str) -> str:
        source_parts = [
            part for part in source_record_key.replace("\\", "/").strip("/").split("/") if part
        ]
        if not source_parts:
            raise ValueError("Không xác định được hồ sơ cần sao chép.")
        existing_keys = {
            record["record_key"]
            for record, _index in self._iter_system_records(project_id)
        }
        new_parts = self._next_record_parts(source_parts, existing_keys)
        self.add_manual_record(project_id, new_parts)
        new_record_key = "/".join(new_parts)

        workflow = self.db.get_record_workflow(project_id, source_record_key)
        if workflow.get("id") is not None:
            self.db.save_record_workflow(
                project_id=project_id,
                record_key=new_record_key,
                scanner_id=workflow.get("scanner_id"),
                scan_date=workflow.get("scan_date", ""),
                checker_id=workflow.get("checker_id"),
                check_date=workflow.get("check_date", ""),
                check_pages=workflow.get("check_pages", 0),
                check_files=workflow.get("check_files", 0),
                record_status=workflow.get("record_status", "NOT_STARTED"),
                notes=workflow.get("notes", ""),
                paper_statuses=[
                    {
                        "paper_format_id": paper["paper_format_id"],
                        "scanner_id": paper.get("scanner_id"),
                        "scan_date": paper.get("scan_date", ""),
                        "scan_status": paper["scan_status"],
                        "scan_pages": str(paper["scan_pages"]),
                        "scan_files": str(paper.get("scan_files", 0)),
                        "check_pages": str(paper["check_pages"]),
                        "notes": paper["notes"],
                    }
                    for paper in workflow["paper_statuses"]
                ],
            )
        return new_record_key

    def _iter_system_records(self, project_id: int):
        offset = 0
        limit = 500
        while True:
            records, total = self.db.list_system_records_page(
                project_id, limit=limit, offset=offset
            )
            for index, record in enumerate(records, start=offset):
                yield record, index
            offset += len(records)
            if offset >= total or not records:
                break

    @staticmethod
    def _next_record_parts(source_parts: list[str], existing_keys: set[str]) -> list[str]:
        prefix = source_parts[:-1]
        last = source_parts[-1]
        if last.isdigit():
            width = len(last)
            number = int(last)
            while True:
                number += 1
                candidate = [*prefix, str(number).zfill(width)]
                if "/".join(candidate) not in existing_keys:
                    return candidate
        suffix = 1
        while True:
            candidate_last = f"{last}-copy" if suffix == 1 else f"{last}-copy-{suffix}"
            candidate = [*prefix, candidate_last]
            if "/".join(candidate) not in existing_keys:
                return candidate
            suffix += 1
