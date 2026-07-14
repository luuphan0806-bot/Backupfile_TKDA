from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from scan_backup_manager.db import MAX_JOBS_PER_PERSON_PER_DAY, Database
from scan_backup_manager.models import JobType, Personnel, Project, ProjectTask
from scan_backup_manager.reports import ReportService


def _project_person(db: Database, tmp_path: Path) -> tuple[int, int]:
    pid = db.create_project(
        Project(
            None,
            "DEMO",
            "Demo",
            str(tmp_path / "b"),
            str(tmp_path / "s"),
            str(tmp_path / "c"),
            str(tmp_path / "r"),
        )
    )
    per = db.save_personnel(Personnel(None, pid, "NV001", "Nguyễn Thị Thương", "Scan"))
    return pid, per


def _task(
    db: Database,
    pid: int,
    per: int,
    code: str,
    day: str,
    *,
    status: str = "COMPLETED",
    record: str = "r",
    title: str = "Scan A3",
) -> None:
    db.save_task(
        ProjectTask(
            None,
            pid,
            code,
            title,
            "",
            per,
            "",
            "NORMAL",
            status,
            record_key=f"{record}/{code}",
            task_kind="SCAN",
            work_date=day,
        )
    )


def test_four_distinct_job_types_hard_limit(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    pid, per = _project_person(db, tmp_path)
    day = date.today().isoformat()
    # 4 *different* job types fill the timesheet's 4 slots.
    for i in range(MAX_JOBS_PER_PERSON_PER_DAY):
        _task(db, pid, per, f"T{i}", day, status="NEW", title=f"Job {i}")
    # A 5th distinct job type is blocked.
    with pytest.raises(ValueError, match="tối đa"):
        _task(db, pid, per, "T5", day, status="NEW", title="Job 5")
    # Repeating an existing job type is unlimited (many records, one slot).
    _task(db, pid, per, "T6", day, status="NEW", title="Job 0")
    _task(db, pid, per, "T7", day, status="NEW", title="Job 0")
    # Editing an existing task is always allowed.
    _task(db, pid, per, "T0", day, status="IN_PROGRESS", title="Job 0")
    # Another work day is independent.
    _task(db, pid, per, "T_OTHER", "2000-01-01", status="NEW", title="Job X")


def test_set_attendance_details_computes_hours_and_validates_type(
    tmp_path: Path,
) -> None:
    db = Database(tmp_path / "app.sqlite3")
    pid, per = _project_person(db, tmp_path)
    day = date.today().isoformat()
    _task(db, pid, per, "T1", day)
    entry = db.list_attendance_entries(pid, day, day)[0]

    db.set_attendance_details(
        int(entry["id"]),
        attendance_type="CC",
        start_time="07:30",
        finish_time="11:30",
        quantity=201,
    )
    updated = db.list_attendance_entries(pid, day, day)[0]
    assert updated["attendance_type"] == "CC"
    assert updated["work_hours"] == 4.0
    assert updated["quantity"] == 201

    with pytest.raises(ValueError, match="Loại chấm công"):
        db.set_attendance_details(int(entry["id"]), attendance_type="ZZ")
    with pytest.raises(ValueError, match="HH:MM"):
        db.set_attendance_details(int(entry["id"]), start_time="25:99")
    with pytest.raises(ValueError, match="không được âm"):
        db.set_attendance_details(int(entry["id"]), work_hours=-1)

    # Lunch break 12:00–13:00 is deducted from any shift that spans it.
    assert db.work_hours_between("07:30", "17:30") == 9.0  # 10h span − 1h lunch
    assert db.work_hours_between("07:30", "11:30") == 4.0  # ends before lunch
    assert db.work_hours_between("11:30", "13:30") == 1.0  # 2h span − 1h lunch
    assert db.work_hours_between("13:00", "17:30") == 4.5  # starts after lunch
    assert db.work_hours_between("10:99", "17:30") == 0.0
    assert db.work_hours_between("", "17:30") == 0.0
    assert db.work_hours_between("18:00", "17:30") == 0.0
    with pytest.raises(ValueError, match="HH:MM"):
        db.set_attendance_details(int(entry["id"]), start_time="10:99")


def test_export_mausham_cong_matches_template_layout(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    pid, per = _project_person(db, tmp_path)
    per2 = db.save_personnel(Personnel(None, pid, "NV002", "Lê Bảo Trân", "Scan"))
    day = date.today().isoformat()
    _task(db, pid, per, "T1", day, title="Scan A3")
    _task(db, pid, per2, "T2", day, title="Scan A3")
    for entry in db.list_attendance_entries(pid, day, day):
        db.set_attendance_details(
            int(entry["id"]),
            attendance_type="CC",
            start_time="07:30",
            finish_time="15:30",
            quantity=201,
        )
        db.approve_attendance_entry(int(entry["id"]), override_reason="test")

    out = ReportService(db).export_mausham_cong(pid, day, day, tmp_path / "out")
    ws = load_workbook(out)[day]

    merged = {str(m) for m in ws.merged_cells.ranges}
    assert {"A1:C2", "D1:G1", "D2:G2", "A4:A7", "B4:B7", "A8:A11", "B8:B11"} <= merged
    assert ws["A3"].value == "Mã NV"
    assert ws["B3"].value == "Họ Và Tên"
    assert ws["E3"].value == "Loại Chấm Công/Năng Suất"
    assert ws["G3"].value == "Khối lượng hoàn thành"
    assert ws["A4"].value == "NV002"
    assert ws["A8"].value == "NV001"
    assert ws["C4"].value == "Công việc 1"
    assert ws["C7"].value == "Công việc 4"
    assert ws["D4"].value == 7  # 07:30–15:30 = 8h span − 1h lunch
    assert ws["E4"].value == "CC"
    assert ws["G4"].value == 201


def test_export_mausham_cong_empty_range_still_produces_sheet(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    pid, _per = _project_person(db, tmp_path)
    out = ReportService(db).export_mausham_cong(
        pid, "2030-01-01", "2030-01-01", tmp_path / "out"
    )
    ws = load_workbook(out)["2030-01-01"]
    assert ws["A3"].value == "Mã NV"


def test_suggested_attendance_quantity_scan_and_check(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    pid, _per = _project_person(db, tmp_path)
    bid = db.upsert_backup_file(
        project_id=pid, client_code="C", source_path="s.pdf", project_code="DEMO",
        relative_project_path="2024/DOC/A1/s.pdf", dest_path="d.pdf", file_size=1,
        source_mtime="2026-01-01T00:00:00", status="HASH_PENDING",
    )
    db.save_backup_file_paper_sizes(bid, {"A3": 201})
    assert db.suggested_attendance_quantity(pid, "2024/DOC/A1", "SCAN") == 201

    db.save_record_workflow(
        project_id=pid, record_key="2024/DOC/A1", scanner_id=None, scan_date="",
        checker_id=None, check_date="", check_pages=88, check_files=3,
        record_status="COMPLETED", notes="", paper_statuses=[],
    )
    assert db.suggested_attendance_quantity(pid, "2024/DOC/A1", "CHECK") == 88
    assert db.suggested_attendance_quantity(pid, "9999/X", "SCAN") == 0


def test_job_limit_is_configurable(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    pid, per = _project_person(db, tmp_path)
    day = date.today().isoformat()
    db.set_setting("max_jobs_per_person_per_day", "2")
    _task(db, pid, per, "A", day, status="NEW", title="Job A")
    _task(db, pid, per, "B", day, status="NEW", title="Job B")
    with pytest.raises(ValueError, match="tối đa 2"):
        _task(db, pid, per, "C", day, status="NEW", title="Job C")
    db.set_setting("max_jobs_per_person_per_day", "0")  # unlimited
    for i in range(5):
        _task(db, pid, per, f"U{i}", day, status="NEW", title=f"Job U{i}")


def test_export_groups_same_job_type_and_sums_volume(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    pid, per = _project_person(db, tmp_path)
    day = date.today().isoformat()
    # Two records of the SAME job type roll into one timesheet slot, volumes summed.
    _task(db, pid, per, "T1", day, title="Scan A3", record="HS1")
    _task(db, pid, per, "T2", day, title="Scan A3", record="HS2")
    # A different job type gets its own slot.
    _task(db, pid, per, "T3", day, title="Scan A0", record="HS3")
    entries = db.list_attendance_entries(pid, day, day)
    for entry in entries:
        qty = 100 if entry["job_title"] == "Scan A3" else 50
        db.set_attendance_details(
            int(entry["id"]), attendance_type="CC",
            start_time="07:30", finish_time="17:30", quantity=qty,
        )
        db.approve_attendance_entry(int(entry["id"]), override_reason="t")
    ws = load_workbook(ReportService(db).export_mausham_cong(pid, day, day, tmp_path / "out"))[day]
    slots = {ws[f"F{r}"].value: ws[f"G{r}"].value for r in range(4, 8)}
    assert slots["Scan A3"] == 200  # 100 + 100 summed across two records
    assert slots["Scan A0"] == 50


def test_export_groups_by_job_type_when_legacy_content_differs(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    pid, per = _project_person(db, tmp_path)
    day = date.today().isoformat()
    _task(db, pid, per, "T1", day, title="Scan A3", record="HS1")
    _task(db, pid, per, "T2", day, title="Scan A3", record="HS2")
    entries = db.list_attendance_entries(pid, day, day)
    for index, entry in enumerate(entries, start=1):
        db.set_attendance_details(
            int(entry["id"]),
            attendance_type="CC",
            start_time="07:30",
            finish_time="17:30",
            job_content=f"Chi tiết hồ sơ {index}",
            quantity=100,
        )
        db.approve_attendance_entry(int(entry["id"]), override_reason="t")

    ws = load_workbook(ReportService(db).export_mausham_cong(pid, day, day, tmp_path / "out"))[day]
    populated = [(ws[f"F{row}"].value, ws[f"G{row}"].value) for row in range(4, 8) if ws[f"F{row}"].value]
    assert populated == [("Chi tiết hồ sơ 1", 200)]


def test_off_app_job_type_persists(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    pid, _per = _project_person(db, tmp_path)
    db.save_job_type(JobType(None, pid, "MANUAL", "Việc tay ngoài app", True, 60, "SCAN", True))
    saved = next(j for j in db.list_job_types(pid) if j.job_code == "MANUAL")
    assert saved.off_app is True
    # Default job types stay app-tracked.
    scan = next(j for j in db.list_job_types(pid) if j.job_code == "SCAN_A4")
    assert scan.off_app is False


def test_completed_off_app_job_can_be_approved_without_backup(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    pid, per = _project_person(db, tmp_path)
    day = date.today().isoformat()
    db.save_job_type(JobType(None, pid, "MANUAL", "Việc tay ngoài app", True, 60, "SCAN", True))
    _task(db, pid, per, "T1", day, title="Việc tay ngoài app")
    entry = db.list_attendance_entries(pid, day, day)[0]

    db.approve_attendance_entry(int(entry["id"]))

    approved = db.list_attendance_entries(pid, day, day)[0]
    assert approved["status"] == "APPROVED"


def test_export_mausham_uses_leader_details(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    pid, per = _project_person(db, tmp_path)
    day = date.today().isoformat()
    _task(db, pid, per, "T1", day, title="Scan A3")
    entry = db.list_attendance_entries(pid, day, day)[0]
    db.set_attendance_details(
        int(entry["id"]), attendance_type="CC.OT",
        start_time="07:30", finish_time="17:30", quantity=201,
    )
    db.approve_attendance_entry(int(entry["id"]), override_reason="t")
    ws = load_workbook(ReportService(db).export_mausham_cong(pid, day, day, tmp_path / "out"))[day]
    assert ws["D4"].value == 9  # 07:30 -> 17:30 = 10h span − 1h lunch
    assert ws["E4"].value == "CC.OT"
    assert ws["G4"].value == 201
