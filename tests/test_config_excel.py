from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from scan_backup_manager.config_excel import ConfigExcelService
from scan_backup_manager.db import Database
from scan_backup_manager.models import Project


def _project(db: Database, tmp_path: Path) -> int:
    return db.create_project(
        Project(
            None,
            "PROJECT_ALPHA",
            "Project Alpha",
            str(tmp_path / "backup"),
            str(tmp_path / "staging"),
            str(tmp_path / "conflicts"),
            str(tmp_path / "reports"),
        )
    )


def test_client_excel_import_and_export(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_id = _project(db, tmp_path)
    service = ConfigExcelService(db)
    source = tmp_path / "clients.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["client_code", "share_path", "enabled", "notes"])
    sheet.append(["scan01", r"\\scan01\share", 1, "Máy scan 01"])
    workbook.save(source)

    count = service.import_clients(project_id, source)
    exported = service.export_clients(project_id, tmp_path / "reports")
    rows = db.list_clients(project_id)

    assert count == 1
    assert rows[0].client_code == "SCAN01"
    assert rows[0].share_path == r"\\scan01\share"
    exported_book = load_workbook(exported, read_only=True)
    assert exported_book.active.max_row == 2


def test_client_template_uses_import_headers(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    service = ConfigExcelService(db)

    template = service.export_client_template(tmp_path)
    sheet = load_workbook(template, read_only=True).active

    assert [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))] == [
        "client_code", "share_path", "enabled", "notes"
    ]


def test_personnel_excel_import_and_export_sets_pin(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_id = _project(db, tmp_path)
    service = ConfigExcelService(db)
    source = tmp_path / "personnel.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["personnel_code", "full_name", "role_name", "enabled", "pin"])
    sheet.append(["nv01", "Nguyễn Văn A", "Scanner", 1, "123456"])
    workbook.save(source)

    count = service.import_personnel(project_id, source)
    exported = service.export_personnel(project_id, tmp_path / "reports")
    people = db.list_personnel(project_id)

    assert count == 1
    assert people[0].personnel_code == "NV01"
    assert db.verify_personnel_pin("PROJECT_ALPHA", "NV01", "123456")
    exported_book = load_workbook(exported, read_only=True)
    assert exported_book.active.max_row == 2


def test_personnel_template_uses_import_headers(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    service = ConfigExcelService(db)

    template = service.export_personnel_template(tmp_path)
    sheet = load_workbook(template, read_only=True).active

    assert [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))] == [
        "personnel_code", "full_name", "role_name", "enabled", "pin"
    ]
