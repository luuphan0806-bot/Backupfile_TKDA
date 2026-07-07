from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from scan_backup_manager.db import Database
from scan_backup_manager.mapfile import MapfileService
from scan_backup_manager.models import DirectoryLevel, Project
from scan_backup_manager.reports import ReportService


def make_project(db: Database, tmp_path: Path) -> int:
    return db.create_project(
        Project(
            None,
            "CSDL_SOHOA_A",
            "Demo",
            str(tmp_path / "backup"),
            str(tmp_path / "staging"),
            str(tmp_path / "conflict_archive"),
            str(tmp_path / "reports"),
        )
    )


def test_mapfile_import_and_reconcile_missing(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_id = make_project(db, tmp_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["project", "year", "case_type", "case_number", "file_name"])
    sheet.append(["CSDL_SOHOA_A", "2023", "HS", "123", "1.pdf"])
    mapfile_path = tmp_path / "mapfile.xlsx"
    workbook.save(mapfile_path)

    import_id = MapfileService(db).import_excel(project_id, mapfile_path)
    rows = db.list_mapfile_rows(import_id)

    assert len(rows) == 1
    assert rows[0]["status"] == "MISSING"
    assert rows[0]["expected_relative_path"] == str(Path("CSDL_SOHOA_A") / "2023" / "HS" / "123" / "1.pdf")


def test_mapfile_import_uses_configured_directory_level_columns(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_id = make_project(db, tmp_path)
    db.save_directory_levels(
        project_id,
        [
            DirectoryLevel(None, project_id, 1, "Nhân sự", "TEXT", []),
            DirectoryLevel(None, project_id, 2, "Ngày", "TEXT", []),
            DirectoryLevel(None, project_id, 3, "Công đoạn", "ENUM", ["SCAN"]),
            DirectoryLevel(None, project_id, 4, "Năm", "YEAR4", []),
            DirectoryLevel(None, project_id, 5, "Loại hồ sơ", "ENUM", ["HS"]),
            DirectoryLevel(None, project_id, 6, "Mã hồ sơ", "TEXT", []),
        ],
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["project", "Nhân sự", "Ngày", "Công đoạn", "Năm", "Loại hồ sơ", "Mã hồ sơ", "file_name"])
    sheet.append(["CSDL_SOHOA_A", "NGUYENVANA", "24052026", "SCAN", "2023", "HS", "123", "1.pdf"])
    mapfile_path = tmp_path / "dynamic_mapfile.xlsx"
    workbook.save(mapfile_path)

    import_id = MapfileService(db).import_excel(project_id, mapfile_path)
    row = db.list_mapfile_rows(import_id)[0]

    assert row["expected_relative_path"] == str(
        Path("CSDL_SOHOA_A") / "NGUYENVANA" / "24052026" / "SCAN" / "2023" / "HS" / "123" / "1.pdf"
    )


def test_mapfile_done_flag_survives_reconcile_and_reimport(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_id = make_project(db, tmp_path)
    service = MapfileService(db)

    def build_workbook(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["project", "year", "case_type", "case_number", "file_name"])
        sheet.append(["CSDL_SOHOA_A", "2023", "HS", "123", "1.pdf"])
        workbook.save(path)

    first_path = tmp_path / "mapfile1.xlsx"
    build_workbook(first_path)
    first_import_id = service.import_excel(project_id, first_path)
    first_row = db.list_mapfile_rows(first_import_id)[0]
    db.mark_mapfile_row_done(first_row["id"], None)

    # reconcile() (auto-triggered by import) must never clear the Done flag it
    # does not own.
    service.reconcile(project_id, first_import_id)
    assert db.get_mapfile_row(first_row["id"])["is_done"] == 1

    # Re-importing the same mapfile creates a brand-new import_id/row, but the
    # Done flag for the same expected_relative_path should carry over.
    second_path = tmp_path / "mapfile2.xlsx"
    build_workbook(second_path)
    second_import_id = service.import_excel(project_id, second_path)
    assert second_import_id != first_import_id
    second_row = db.list_mapfile_rows(second_import_id)[0]
    assert second_row["is_done"] == 1


def test_mapfile_row_cell_update_rebuilds_expected_path_and_reconciles(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_id = make_project(db, tmp_path)
    service = MapfileService(db)

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["project", "year", "case_type", "case_number", "file_name"])
    sheet.append(["CSDL_SOHOA_A", "2023", "HS", "123", "1.pdf"])
    mapfile_path = tmp_path / "mapfile.xlsx"
    workbook.save(mapfile_path)
    import_id = service.import_excel(project_id, mapfile_path)
    row = db.list_mapfile_rows(import_id)[0]

    db.upsert_backup_file(
        project_id=project_id,
        client_code="SCAN01",
        source_path=r"\\SCAN01\share\CSDL_SOHOA_A\2023\HS\456\1.pdf",
        project_code="CSDL_SOHOA_A",
        relative_project_path="2023/HS/456/1.pdf",
        dest_path=r"D:\backup\CSDL_SOHOA_A\2023\HS\456\1.pdf",
        file_size=100,
        source_mtime="2026-07-07T00:00:00+00:00",
        status="LOCKED",
    )

    status = service.update_row_cell(project_id, row["id"], "case_number", "456")
    updated = db.get_mapfile_row(row["id"])

    assert status == "MATCHED"
    assert json.loads(updated["raw_json"])["case_number"] == "456"
    assert updated["expected_relative_path"] == str(
        Path("CSDL_SOHOA_A") / "2023" / "HS" / "456" / "1.pdf"
    )
    assert updated["record_key"] == "2023/HS/456"
    assert updated["status"] == "MATCHED"


def test_report_export_creates_excel(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_id = make_project(db, tmp_path)

    report_path = ReportService(db).export_daily_report(project_id)

    assert report_path.exists()
    assert report_path.suffix == ".xlsx"
    assert report_path.parent == tmp_path / "reports"
