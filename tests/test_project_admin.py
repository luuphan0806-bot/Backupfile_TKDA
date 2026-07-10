from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from scan_backup_manager.db import DEFAULT_ADMIN_PASSWORD, Database
from scan_backup_manager.mapfile import MapfileService
from scan_backup_manager.models import Client, DirectoryLevel, JobType, Personnel, Project, ProjectTask
from scan_backup_manager.reports import ReportService


def configure_project(db: Database, tmp_path: Path) -> Project:
    project_id = db.create_project(
        Project(
            None,
            "PROJECT_ALPHA",
            "Project Alpha",
            str(tmp_path / "backup"),
            str(tmp_path / "staging"),
            str(tmp_path / "conflicts"),
            str(tmp_path / "reports"),
        )
    )
    db.save_directory_levels(
        project_id,
        [
            DirectoryLevel(None, project_id, 1, "Year", "YEAR4", []),
            DirectoryLevel(None, project_id, 2, "Batch", "INTEGER", []),
            DirectoryLevel(None, project_id, 3, "Category", "ENUM", ["DOC", "INVOICE"]),
            DirectoryLevel(None, project_id, 4, "Record", "TEXT", []),
        ],
    )
    project = db.get_project(project_id)
    assert project is not None
    return project


def test_default_admin_password_must_be_changed(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")

    assert db.verify_admin_password(DEFAULT_ADMIN_PASSWORD)
    assert db.admin_must_change_password()
    assert not db.verify_admin_password("wrong")

    db.change_admin_password(DEFAULT_ADMIN_PASSWORD, "A-new-password")

    assert not db.verify_admin_password(DEFAULT_ADMIN_PASSWORD)
    assert db.verify_admin_password("A-new-password")
    assert not db.admin_must_change_password()


def test_single_project_and_dynamic_levels(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project = configure_project(db, tmp_path)

    assert project.project_code == "PROJECT_ALPHA"
    assert [level.validation_type for level in db.list_directory_levels(project.id)] == [
        "YEAR4", "INTEGER", "ENUM", "TEXT"
    ]


def test_duplicate_project_code_shows_validation_error(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project = configure_project(db, tmp_path)

    with pytest.raises(ValueError, match="Mã dự án 'PROJECT_ALPHA' đã tồn tại"):
        db.create_project(
            Project(
                None,
                "project_alpha",
                "Duplicate",
                str(tmp_path / "backup_2"),
                str(tmp_path / "staging_2"),
                str(tmp_path / "conflicts_2"),
                str(tmp_path / "reports_2"),
            )
        )

    other_id = db.create_project(
        Project(
            None,
            "PROJECT_BETA",
            "Project Beta",
            str(tmp_path / "backup_b"),
            str(tmp_path / "staging_b"),
            str(tmp_path / "conflicts_b"),
            str(tmp_path / "reports_b"),
        )
    )
    other = db.get_project(other_id)
    assert other is not None
    other.project_code = project.project_code

    with pytest.raises(ValueError, match="Mã dự án 'PROJECT_ALPHA' đã tồn tại"):
        db.save_project(other)


def test_directory_level_allowed_values_preserve_display_text(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project = configure_project(db, tmp_path)

    db.save_directory_levels(
        project.id or 0,
        [
            DirectoryLevel(None, project.id or 0, 1, "Nhân sự", "TEXT", ["Nguyễn Văn A"]),
            DirectoryLevel(None, project.id or 0, 2, "Loại hồ sơ", "ENUM", ["Hồ sơ"]),
        ],
    )

    levels = db.list_directory_levels(project.id or 0)
    assert levels[0].allowed_values == ["Nguyễn Văn A"]
    assert levels[1].allowed_values == ["Hồ sơ"]


def test_directory_levels_store_mapfile_display_settings(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project = configure_project(db, tmp_path)
    db.save_directory_levels(
        project.id or 0,
        [
            DirectoryLevel(None, project.id or 0, 1, "Năm", "YEAR4", ["2024"], True, 2),
            DirectoryLevel(None, project.id or 0, 2, "Loại hồ sơ", "ENUM", ["DOC"], False, 1),
        ],
    )

    levels = db.list_directory_levels(project.id or 0)

    assert levels[0].show_in_mapfile is True
    assert levels[0].mapfile_position == 2
    assert levels[1].show_in_mapfile is False
    assert levels[1].mapfile_position == 1


def test_directory_levels_store_catalog_requirement(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project = configure_project(db, tmp_path)
    db.save_directory_levels(
        project.id or 0,
        [
            DirectoryLevel(None, project.id or 0, 1, "Loại hồ sơ", "ENUM", ["DOC"], True, 1, True),
        ],
    )

    levels = db.list_directory_levels(project.id or 0)

    assert levels[0].require_catalog_selection is True


def test_required_catalog_level_needs_values(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project = configure_project(db, tmp_path)

    with pytest.raises(ValueError, match="bắt buộc chọn từ danh mục"):
        db.save_directory_levels(
            project.id or 0,
            [
                DirectoryLevel(None, project.id or 0, 1, "Loại hồ sơ", "ENUM", [], True, 1, True),
            ],
        )


def test_enum_directory_level_can_start_without_catalog_values(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project = configure_project(db, tmp_path)

    db.save_directory_levels(
        project.id or 0,
        [DirectoryLevel(None, project.id or 0, 1, "Loại hồ sơ", "ENUM", [])],
    )

    levels = db.list_directory_levels(project.id or 0)
    assert levels[0].display_name == "Loại hồ sơ"
    assert levels[0].allowed_values == []


def test_multiple_projects_do_not_leak_data(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_a = configure_project(db, tmp_path)
    project_b_id = db.create_project(
        Project(
            None,
            "PROJECT_BETA",
            "Project Beta",
            str(tmp_path / "backup_b"),
            str(tmp_path / "staging_b"),
            str(tmp_path / "conflicts_b"),
            str(tmp_path / "reports_b"),
        )
    )
    db.save_directory_levels(
        project_b_id, [DirectoryLevel(None, project_b_id, 1, "Year", "YEAR4", [])]
    )

    # Same workstation code and personnel code are allowed in both projects,
    # since uniqueness is scoped per project_id.
    db.save_personnel(Personnel(None, project_a.id or 0, "NV01", "Nguyen Van A", "Scanner", True))
    db.save_personnel(Personnel(None, project_b_id, "NV01", "Tran Van B", "Scanner", True))

    assert [person.full_name for person in db.list_personnel(project_a.id or 0)] == ["Nguyen Van A"]
    assert [person.full_name for person in db.list_personnel(project_b_id)] == ["Tran Van B"]
    assert len(db.list_projects()) == 2


def test_inactive_personnel_cannot_receive_new_task(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project = configure_project(db, tmp_path)
    personnel_id = db.save_personnel(
        Personnel(None, project.id or 0, "NV01", "Nguyen Van A", "Scanner", False)
    )

    with pytest.raises(ValueError, match="active personnel"):
        db.save_task(
            ProjectTask(
                None, project.id or 0, "TASK01", "Scan batch", "",
                personnel_id, "2026-07-10",
            )
        )


def test_assigned_personnel_is_deactivated_instead_of_deleted(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project = configure_project(db, tmp_path)
    personnel_id = db.save_personnel(
        Personnel(None, project.id or 0, "NV01", "Nguyen Van A", "Scanner", True)
    )
    db.save_task(
        ProjectTask(
            None, project.id or 0, "TASK01", "Scan batch", "",
            personnel_id, "2026-07-10",
        )
    )

    db.delete_personnel(personnel_id)

    assert db.list_personnel(project.id or 0)[0].enabled is False
    assert len(db.list_tasks(project.id or 0)) == 1


def test_legacy_database_is_backed_up_and_reset(tmp_path: Path) -> None:
    db_path = tmp_path / "legacy.sqlite3"
    conn = sqlite3.connect(db_path)
    try:
        conn.execute("CREATE TABLE clients(id INTEGER PRIMARY KEY, client_code TEXT)")
        conn.execute("INSERT INTO clients(client_code) VALUES('OLD')")
        conn.commit()
    finally:
        conn.close()

    db = Database(db_path)

    assert db.migration_backup_path is not None
    assert db.migration_backup_path.exists()
    assert db.list_projects() == []


def test_v2_database_migrates_in_place_without_data_loss(tmp_path: Path) -> None:
    db_path = tmp_path / "legacy_v2.sqlite3"
    conn = sqlite3.connect(db_path)
    try:
        conn.executescript(
            """
            CREATE TABLE app_meta(key TEXT PRIMARY KEY, value TEXT NOT NULL);
            CREATE TABLE admin_auth(
                id INTEGER PRIMARY KEY CHECK(id = 1),
                salt TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                must_change_password INTEGER NOT NULL DEFAULT 1,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE projects(
                id INTEGER PRIMARY KEY,
                singleton INTEGER NOT NULL DEFAULT 1 UNIQUE CHECK(singleton = 1),
                project_code TEXT NOT NULL UNIQUE,
                display_name TEXT NOT NULL,
                backup_root TEXT NOT NULL,
                staging_dir TEXT NOT NULL,
                conflict_archive_dir TEXT NOT NULL,
                reports_dir TEXT NOT NULL,
                enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE clients(
                id INTEGER PRIMARY KEY,
                project_id INTEGER NOT NULL,
                client_code TEXT NOT NULL,
                share_path TEXT NOT NULL,
                staff_name TEXT NOT NULL DEFAULT '',
                enabled INTEGER NOT NULL DEFAULT 1,
                notes TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(project_id, client_code)
            );
            CREATE TABLE settings(key TEXT PRIMARY KEY, value TEXT NOT NULL);
            CREATE TABLE backup_files(
                id INTEGER PRIMARY KEY,
                project_id INTEGER NOT NULL,
                client_code TEXT NOT NULL,
                source_path TEXT NOT NULL,
                project_code TEXT NOT NULL,
                relative_project_path TEXT NOT NULL,
                dest_path TEXT NOT NULL,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(project_id, client_code, source_path)
            );
            CREATE TABLE mapfile_profiles(
                id INTEGER PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                sheet_name TEXT NOT NULL DEFAULT '',
                project_column TEXT NOT NULL,
                year_column TEXT NOT NULL,
                case_type_column TEXT NOT NULL,
                case_number_column TEXT NOT NULL,
                file_name_column TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            """
        )
        conn.execute("INSERT INTO app_meta(key, value) VALUES('schema_version', '2')")
        conn.execute(
            """
            INSERT INTO projects(
                id, singleton, project_code, display_name, backup_root, staging_dir,
                conflict_archive_dir, reports_dir, enabled, created_at, updated_at
            ) VALUES(1, 1, 'LEGACY_PROJECT', 'Legacy Project', 'D:\\BACKUP', 'data/staging',
                'data/conflict_archive', 'data/reports', 1, '2026-01-01T00:00:00+00:00',
                '2026-01-01T00:00:00+00:00')
            """
        )
        conn.execute(
            """
            INSERT INTO clients(
                id, project_id, client_code, share_path, staff_name, enabled, notes,
                created_at, updated_at
            ) VALUES(1, 1, 'SCAN01', 'C:\\share', '', 1, '', '2026-01-01T00:00:00+00:00',
                '2026-01-01T00:00:00+00:00')
            """
        )
        conn.execute(
            """
            INSERT INTO backup_files(
                id, project_id, client_code, source_path, project_code,
                relative_project_path, dest_path, status, created_at
            ) VALUES(1, 1, 'SCAN01', 'C:\\share\\file.pdf', 'LEGACY_PROJECT',
                '2026/HS/1/file.pdf', 'D:\\BACKUP\\LEGACY_PROJECT\\2026\\HS\\1\\file.pdf',
                'LOCKED', '2026-01-01T00:00:00+00:00')
            """
        )
        conn.execute(
            """
            INSERT INTO mapfile_profiles(
                name, sheet_name, project_column, year_column, case_type_column,
                case_number_column, file_name_column, created_at, updated_at
            ) VALUES('Default', '', 'project', 'year', 'case_type', 'case_number', 'file_name',
                '2026-01-01T00:00:00+00:00', '2026-01-01T00:00:00+00:00')
            """
        )
        conn.execute("INSERT INTO settings(key, value) VALUES('poll_interval_seconds', '120')")
        conn.commit()
    finally:
        conn.close()

    db = Database(db_path)

    assert db.migration_backup_path is not None
    assert db.migration_backup_path.exists()

    projects = db.list_projects()
    assert len(projects) == 1
    project = projects[0]
    assert project.project_code == "LEGACY_PROJECT"

    clients = db.list_clients(project.id or 0)
    assert [client.client_code for client in clients] == ["SCAN01"]

    backup_files = db.list_backup_files(project.id or 0)
    assert len(backup_files) == 1
    assert backup_files[0]["status"] == "LOCKED"

    profile = db.get_mapfile_profile(project.id or 0, "Default")
    assert profile.project_column == "project"

    settings = db.get_project_settings(project.id or 0)
    assert settings.poll_interval_seconds == 120


def test_report_contains_project_personnel_and_tasks(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project = configure_project(db, tmp_path)
    personnel_id = db.save_personnel(
        Personnel(None, project.id or 0, "NV01", "Nguyen Van A", "Scanner", True)
    )
    db.save_task(
        ProjectTask(
            None, project.id or 0, "TASK01", "Scan batch", "",
            personnel_id, "2026-07-10",
        )
    )

    report = ReportService(db).export_daily_report(project.id or 0)
    workbook = load_workbook(report, read_only=True)

    assert {"Summary", "Personnel", "Tasks"}.issubset(workbook.sheetnames)


def test_project_job_types_can_be_renamed_and_used_for_manual_work(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project = configure_project(db, tmp_path)
    job_types = db.list_job_types(project.id or 0)
    assert [item.display_name for item in job_types] == [
        "Scan A4",
        "Scan A3 (mới)",
        "Scan A3 (cũ)",
        "Scan A0",
        "Check Scan",
    ]
    kinds = {item.job_code: item.job_kind for item in job_types}
    assert kinds == {
        "SCAN_A4": "SCAN",
        "SCAN_A3": "SCAN",
        "SCAN_A3_OLD": "SCAN",
        "SCAN_A0": "SCAN",
        "CHECK": "CHECK",
    }
    scan_a4 = next(item for item in job_types if item.job_code == "SCAN_A4")

    db.save_job_type(
        JobType(
            scan_a4.id,
            project.id or 0,
            "SCAN_A4",
            "Quét hồ sơ A4",
            True,
            scan_a4.sort_order,
            "CHECK",
        )
    )
    renamed = next(item for item in db.list_job_types(project.id or 0) if item.job_code == "SCAN_A4")
    assert renamed.display_name == "Quét hồ sơ A4"
    assert renamed.job_kind == "CHECK"

    share = tmp_path / "share"
    db.save_client(Client(None, project.id or 0, "SCAN01", "", str(share), True))
    personnel_id = db.save_personnel(
        Personnel(None, project.id or 0, "NV01", "Nguyen Van A", "Scanner", True)
    )
    row_id = MapfileService(db).add_manual_record(
        project.id or 0,
        ["2026", "100", "DOC", "A-001"],
        client_code="SCAN01",
    )
    task_id = db.save_task(
        ProjectTask(
            None,
            project.id or 0,
            "SCAN_A4_2026_100_DOC_A_001",
            renamed.display_name,
            f"Dòng mapfile: {row_id}",
            personnel_id,
            "",
        )
    )

    assert task_id > 0
    assert (
        share
        / "CSDL_SOHOA_PROJECT_ALPHA"
        / "Họ tên"
        / datetime.now().strftime("%d-%m-%Y")
        / "Nội dung công việc"
        / "2026"
        / "100"
        / "DOC"
        / "A-001"
    ).is_dir()
