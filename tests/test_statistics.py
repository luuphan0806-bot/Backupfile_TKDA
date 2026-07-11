from pathlib import Path

import pytest
from openpyxl import load_workbook

from scan_backup_manager.db import Database
from scan_backup_manager.models import Personnel, Project, ProjectTask
from scan_backup_manager.reports import ReportService
from scan_backup_manager.statistics import StatisticsService


def test_job_quantity_by_day_groups_by_date_and_job(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_id = db.create_project(
        Project(
            None,
            "DEMO",
            "Demo",
            str(tmp_path / "backup"),
            str(tmp_path / "staging"),
            str(tmp_path / "conflicts"),
            str(tmp_path / "reports"),
        )
    )
    person_id = db.save_personnel(
        Personnel(None, project_id, "NV01", "Nguyen Van A", "Scanner")
    )
    db.save_task(
        ProjectTask(
            None,
            project_id,
            "SCAN_A4_001",
            "Scan A4",
            "Record 001",
            person_id,
            "",
            status="COMPLETED",
            record_key="2025/BAN_VE/001",
            task_kind="SCAN",
        )
    )
    db.save_task(
        ProjectTask(
            None,
            project_id,
            "SCAN_A3_001",
            "Scan A3",
            "Record 001",
            person_id,
            "",
            status="COMPLETED",
            record_key="2025/BAN_VE/001-A3",
            task_kind="SCAN",
        )
    )
    db.save_task(
        ProjectTask(
            None,
            project_id,
            "CHECK_001",
            "Check Scan",
            "Record 001",
            person_id,
            "",
            status="COMPLETED",
            record_key="2025/BAN_VE/001",
            task_kind="CHECK",
        )
    )

    assert StatisticsService(db).job_quantity_by_day(
        project_id, "2000-01-01", "2999-12-31"
    ) == []

    for row in db.list_attendance_entries(project_id, "2000-01-01", "2999-12-31"):
        db.approve_attendance_entry(
            int(row["id"]),
            override_reason="Leader verified sample output for test",
        )

    rows = StatisticsService(db).job_quantity_by_day(
        project_id, "2000-01-01", "2999-12-31"
    )

    by_job = {(row.task_kind, row.job_title): row for row in rows}
    assert by_job[("SCAN", "Scan A4")].quantity == 1
    assert by_job[("SCAN", "Scan A3")].quantity == 1
    assert by_job[("CHECK", "Check Scan")].quantity == 1
    assert by_job[("CHECK", "Check Scan")].completed_count == 1

    details = StatisticsService(db).personnel_daily_job_details(
        project_id, "2000-01-01", "2999-12-31"
    )
    assert [row.sequence_number for row in details] == [1, 2, 3]
    assert [(row.job_title, row.quantity) for row in details] == [
        ("Scan A4", 1),
        ("Scan A3", 1),
        ("Check Scan", 1),
    ]
    assert all(row.started_at for row in details)

    output = ReportService(db).export_attendance_report(
        project_id, "2000-01-01", "2999-12-31", tmp_path / "reports"
    )
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Cham cong", "Tong hop", "San luong tho", "Ngoai le", "Audit chinh sua"]
    attendance = workbook["Cham cong"]
    rows = list(attendance.iter_rows(values_only=True))
    assert rows[0][:5] == (
        "day",
        "personnel_code",
        "full_name",
        "sequence_number",
        "job_title",
    )
    assert len(rows) == 4
    assert {row[4] for row in rows[1:]} == {"Scan A4", "Scan A3", "Check Scan"}


def test_attendance_approval_requires_completion_and_override_reason(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_id = db.create_project(
        Project(
            None,
            "DEMO",
            "Demo",
            str(tmp_path / "backup"),
            str(tmp_path / "staging"),
            str(tmp_path / "conflicts"),
            str(tmp_path / "reports"),
        )
    )
    person_id = db.save_personnel(
        Personnel(None, project_id, "NV01", "Nguyen Van A", "Scanner")
    )
    db.save_task(
        ProjectTask(
            None,
            project_id,
            "SCAN_A4_002",
            "Scan A4",
            "Record 002",
            person_id,
            "",
            record_key="2025/BAN_VE/002",
            task_kind="SCAN",
        )
    )
    entry = db.list_attendance_entries(project_id, "2000-01-01", "2999-12-31")[0]

    with pytest.raises(ValueError, match="chưa được chốt"):
        db.approve_attendance_entry(int(entry["id"]))

    db.complete_open_tasks_for_assignee(project_id, person_id)
    entry = db.list_attendance_entries(project_id, "2000-01-01", "2999-12-31")[0]
    with pytest.raises(ValueError, match="backup"):
        db.approve_attendance_entry(int(entry["id"]))

    db.approve_attendance_entry(
        int(entry["id"]),
        override_reason="Leader accepted manual source evidence",
    )
    approved = db.list_attendance_entries(
        project_id, "2000-01-01", "2999-12-31", statuses=["APPROVED"]
    )
    assert len(approved) == 1
    assert approved[0]["override_reason"] == "Leader accepted manual source evidence"


def _project_with_person(db: Database, tmp_path: Path) -> tuple[int, int]:
    project_id = db.create_project(
        Project(
            None,
            "DEMO",
            "Demo",
            str(tmp_path / "backup"),
            str(tmp_path / "staging"),
            str(tmp_path / "conflicts"),
            str(tmp_path / "reports"),
        )
    )
    person_id = db.save_personnel(
        Personnel(None, project_id, "NV01", "Nguyen Van A", "Scanner")
    )
    return project_id, person_id


def test_reject_attendance_requires_reason_and_is_excluded_from_stats(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_id, person_id = _project_with_person(db, tmp_path)
    db.save_task(
        ProjectTask(
            None, project_id, "SCAN_R", "Scan A4", "Record", person_id, "",
            status="COMPLETED", record_key="2025/BAN_VE/010", task_kind="SCAN",
        )
    )
    entry = db.list_attendance_entries(project_id, "2000-01-01", "2999-12-31")[0]

    with pytest.raises(ValueError, match="lý do"):
        db.reject_attendance_entry(int(entry["id"]))

    db.reject_attendance_entry(int(entry["id"]), reason="Sản lượng không đạt")
    rejected = db.list_attendance_entries(
        project_id, "2000-01-01", "2999-12-31", statuses=["REJECTED"]
    )
    assert len(rejected) == 1
    assert rejected[0]["notes"] == "Sản lượng không đạt"
    # Rejected work never reaches the approved statistics.
    assert StatisticsService(db).job_quantity_by_day(
        project_id, "2000-01-01", "2999-12-31"
    ) == []


def test_override_approval_adjusts_quantity_and_completed(tmp_path: Path) -> None:
    db = Database(tmp_path / "app.sqlite3")
    project_id, person_id = _project_with_person(db, tmp_path)
    db.save_task(
        ProjectTask(
            None, project_id, "SCAN_O", "Scan A4", "Record", person_id, "",
            status="COMPLETED", record_key="2025/BAN_VE/011", task_kind="SCAN",
        )
    )
    entry = db.list_attendance_entries(project_id, "2000-01-01", "2999-12-31")[0]

    db.approve_attendance_entry(
        int(entry["id"]),
        quantity=5,
        completed_count=3,
        override_reason="Gộp sản lượng nhiều file",
    )

    approved = db.list_attendance_entries(
        project_id, "2000-01-01", "2999-12-31", statuses=["APPROVED"]
    )[0]
    assert approved["quantity"] == 5
    assert approved["completed_count"] == 3

    rows = StatisticsService(db).job_quantity_by_day(
        project_id, "2000-01-01", "2999-12-31"
    )
    assert rows[0].quantity == 5
    assert rows[0].completed_count == 3
